import anthropic
from openai import OpenAI
from openai.types.chat.chat_completion_message_param import ChatCompletionMessageParam
from typing import List
import openpyxl
from openpyxl import Workbook, load_workbook
import time
import base64
import os
from dotenv import load_dotenv
from reka.client import Reka
import pandas as pd

# Function to convert image to base64
def get_image_base64(slide):
    image_path = f"Supplementary_Material_2/{slide}.jpg"
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image file not found: {image_path}")
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

# Set up API clients
load_dotenv('api_claude.env')
anthropic_client = anthropic.Anthropic(api_key=os.getenv('API_KEY_CLAUDE'))

load_dotenv('api_gpt4.env')
openai_client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

load_dotenv('api_reka.env')
reka_client = Reka(api_key=os.getenv('REKA_API_KEY'))

model_dict = {
    "GPT-4o": "gpt-4o-2024-05-13",
    "Claude-3.5": "claude-3-5-sonnet-20240620",
    "Claude-3": "claude-3-opus-20240229",
    "Reka-Core": "reka-core-20240501"
}

MAX_TOKENS = 1000
TEMPERATURE = 0.7

# Define prompts
prompt_dictionary = {
    "Default": "Describe this image. Follow this Template: Organ:_____ Tissue of Origin: ____Pathologies: ______  {if applicable, else ""None""}.",
    "Healthy": "Just describe which organ you see but state it looks healthy"
}

# Define slide range and special cases
slide_range = [f"Folie{i}" for i in range(1, 91)]
prompt2_cases = ['Folie47', 'Folie52', 'Folie57', 'Folie62', 'Folie67', 'Folie72', 'Folie77', 'Folie82', 'Folie87']
sequential_list = [('Folie48', 'Folie49'), ('Folie53', 'Folie54'), ('Folie58', 'Folie59'), ('Folie63', 'Folie64'), ('Folie68', 'Folie69'), ('Folie73', 'Folie74'), ('Folie78', 'Folie79'), ('Folie83', 'Folie84'), ('Folie88', 'Folie89')]

# Analysis functions
def analyze_image_claude(slide, prompt, model, previous_messages=None):
    try:
        base64_image = get_image_base64(slide)
        content = [
            {"type": "text", "text": prompt},
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/jpeg",
                    "data": base64_image
                }
            }
        ]
        if previous_messages:
            content = previous_messages + content
        message = anthropic_client.messages.create(
            model=model,
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
            messages=[
                {
                    "role": "user",
                    "content": content
                }
            ]
        )
        return message.content[0].text
    except Exception as e:
        return f"Error analyzing {slide}: {str(e)}"

def analyze_image_gpt4(slide, prompt, model, previous_messages=None):
    try:
        base64_image = get_image_base64(slide)
        messages: List[ChatCompletionMessageParam] = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ]
        if previous_messages:
            messages = previous_messages + messages
        
        response = openai_client.chat.completions.create(
            model=model,
            messages=messages,
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
        )
        
        if response.choices and len(response.choices) > 0:
            return response.choices[0].message.content
        else:
            return "No response generated"
    except Exception as e:
        return f"Error analyzing {slide}: {str(e)}"

def analyze_image_reka(slide, prompt, model, previous_messages=None):
    try:
        base64_image = get_image_base64(slide)
        
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": f"data:image/jpeg;base64,{base64_image}"
                    }
                ]
            }
        ]
        if previous_messages:
            messages = previous_messages + messages
        
        response = reka_client.chat.create(
            messages=messages,
            model=model,
            max_tokens=MAX_TOKENS,
            temperature=TEMPERATURE,
        )
        
        if response.responses and len(response.responses) > 0:
            return response.responses[0].message.content
        else:
            return "No response generated"
    except Exception as e:
        return f"Error analyzing {slide}: {str(e)}"

def get_analysis_function(model_name):
    if model_name.startswith("Claude"):
        return analyze_image_claude
    elif model_name.startswith("GPT"):
        return analyze_image_gpt4
    elif model_name == "Reka-Core":
        return analyze_image_reka
    else:
        raise ValueError(f"Unknown model: {model_name}")

def append_to_excel(results, filename, sheet_name="Visual Prompt Experiment"):
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        headers = ["Slide", "Prompt", "Model Name", "Model ID", "Result 1", "Result 2", "Result 3"]
        ws.append(headers)
    else:
        ws = wb[sheet_name]

    for row in results:
        ws.append(row)

    wb.save(filename)

# Main analysis loop
output_file = "visual_prompt_experiment_results.xlsx"
model_list = ["Claude-3.5", "Claude-3", "GPT-4o", "Reka-Core"]

for model in model_list:
    all_results = []
    analysis_function = get_analysis_function(model)
    model_id = model_dict[model]

    for slide in slide_range:
        prompt_variation = "Healthy" if slide in prompt2_cases else "Default"
        prompt = prompt_dictionary[prompt_variation]
        results = []
        
        for execution in range(3):  # 3 executions
            response = analysis_function(slide, prompt, model_id)
            results.append(response)
            time.sleep(1)
        
        all_results.append([slide, prompt, model, model_id] + results)
        
        # Handle sequential slides
        if any(slide == seq[0] for seq in sequential_list):
            next_slide = [seq[1] for seq in sequential_list if seq[0] == slide][0]
            next_results = []
            
            for execution in range(3):  # 3 executions
                previous_messages = [
                    {"type": "text", "text": prompt},
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": get_image_base64(slide)}},
                    {"type": "text", "text": results[execution]}
                ]
                next_response = analysis_function(next_slide, prompt_dictionary["Default"], model_id, previous_messages)
                next_results.append(next_response)
                time.sleep(1)
            
            all_results.append([next_slide, prompt_dictionary["Default"], model, model_id] + next_results)
            
            # Print the conversation for sequential slides to the console
            print(f"\nSequential Analysis for {slide} and {next_slide} using {model}:")
            for i in range(3):
                print(f"Execution {i+1}:")
                print(f"Prompt for {slide}: {prompt}")
                print(f"Response for {slide}: {results[i]}")
                print(f"Prompt for {next_slide}: {prompt_dictionary['Default']}")
                print(f"Response for {next_slide}: {next_results[i]}")
                print("-" * 50)
            
            # Remove the second slide from the main loop if it's in the slide range
            if next_slide in slide_range:
                slide_range.remove(next_slide)

    append_to_excel(all_results, output_file, sheet_name=f"Results_{model}")
    print(f"Analysis completed for {model}. Results appended to {output_file}")

# Create a filtered results sheet
wb = load_workbook(output_file)
for model in model_list:
    ws = wb[f"Results_{model}"]
    ws_filtered = wb.create_sheet(title=f"Filtered_Results_{model}")
    
    # Copy headers
    ws_filtered.append(next(ws.iter_rows(values_only=True)))
    
    # Get the first items of each sequential pair
    first_sequential = [seq[0] for seq in sequential_list]
    
    # Copy rows from the first worksheet to the second, excluding the first sequential items
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] not in first_sequential:
            ws_filtered.append(row)

wb.save(output_file)
print(f"Experiment completed. Results saved to {output_file}")
